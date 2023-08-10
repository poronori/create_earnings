import openpyxl
import configparser
import datetime
from copy import copy

date = '2023/8/1'
name = 'なんかの品名'
price = '1000'
commission = '100'
customer = 'テスト 太郎'
address = '北海道 名古屋市 伏見'
code = 'a99999999'

def edit() : 
    #設定ファイルからエクセルのパスを取得
    inifile = configparser.SafeConfigParser()
    inifile.read('config/config.ini', encoding='utf-8')
    excel_path = inifile.get('DEFAULT', 'ExcelPath')
    
    #現年の販売リストシートを開く
    wb = openpyxl.load_workbook(excel_path)
    current_year = datetime.date.today().year
    ws = wb['販売リスト %s' %current_year]
    
    #一番下の行を取得
    maxRow = ws.max_row
    #max_rowを使うと削除していた行も取得してしまうため、Noneで判定する
    for row in ws.iter_rows():
        if not row[0] or row[0].value is None :
            maxRow = row[0].row - 1
            break
    print(maxRow)
    nextRow = maxRow + 1
    
    #行を挿入
    ws.insert_rows(nextRow)
    
    #書式をコピー
    i = 1
    profit = ''
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == maxRow:
                ws.cell(row = nextRow, column = i).border = copy(cell.border)
                ws.cell(row = nextRow, column = i)._style = copy(cell._style)
                value = cell.value
                #計算式をコピー
                if str(value)[0] == "=":
                    profit = value.replace(str(maxRow), str(nextRow))
                i = i + 1

    #値をセット
    no = ws.cell(row = maxRow, column = 1).value
    ws.cell(row = nextRow, column = 1).value = int(no) + 1 #No
    ws.cell(row = nextRow, column = 2).value = string_to_datetime(date)    #購入日
    ws.cell(row = nextRow, column = 3).value = name        #品名
    ws.cell(row = nextRow, column = 4).value = int(price)  #商品代金
    ws.cell(row = nextRow, column = 5).value = int(commission)  #販売手数料
    ws.cell(row = nextRow, column = 6).value = ''          #梱包資材１   
    ws.cell(row = nextRow, column = 7).value = '封筒'          #梱包資材２
    ws.cell(row = nextRow, column = 8).value = ''          #送料
    ws.cell(row = nextRow, column = 9).value = profit      #販売利益
    ws.cell(row = nextRow, column = 10).value = customer   #購入者
    ws.cell(row = nextRow, column = 11).value = prefecture_from_address(address)    #住所
    ws.cell(row = nextRow, column = 12).value = code       #コード

    wb.save(excel_path)

#日時を〇月×日にフォーマット
def string_to_datetime(date_string) :
    datetime_array = date_string.split('/')
    year = int(datetime_array[0])
    month = int(datetime_array[1])
    day = int(datetime_array[2])
    dt = datetime.date(year, month, day)
    
    return dt.strftime('%m月%d日')

#住所から県名を抽出
def prefecture_from_address(address) :
    index = address.find('県', 0, 4)
    if index != -1 :
        return address[0:index + 1]
    
    index = address.find('府', 0, 3)
    if index != -1 :
        return address[0:index + 1]
    
    if '北海道' in address :
        return address[0:3]
    
    if '東京都' in address :
        return address[0:3]